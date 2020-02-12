from PIL import Image
import pytesseract
import os
from os.path import join
import datetime
from openpyxl import load_workbook


# Define path for xlsx sheet
XslxPath = r''

# Define path for files ready for processing
path = r''

# Define Paths for processed files

ProcessedPdfPath = ''
ErrorPdfPath = ''
ProcessedImagePath = ''
ErrorImagePath = ''


# load xlsx workbook for logging valid processed images
wb = load_workbook(XslxPath)
ws = wb.active
lastRow = ws.max_row + 1

# create list containing each file in the defined path
fileList = [x for x in os.listdir(path) if os.path.isfile(join(path, x))]

# Optional. Set the pytesseract source folder if pytesseract is not in the PATH variable
pytesseract.pytesseract.tesseract_cmd = r''

# loop through the files in the defined file path

for files in fileList:

    currentDate = datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")

    img = Image.open(path + files)

# optional. crop image to simplify the OCR process
    width, height = img.size

    left = 2300
    top = height / 18
    right = 3000
    bottom = height / 14.5

    img1 = img.crop((left, top, right, bottom))
    img1.show()

    # start Tesseract process
    try:

        result = pytesseract.image_to_string(img1,  config="-c tessedit_char_whitelist=0123456789 --psm 7")

        # check if result is 8 characters long and start with int 2
        if len(result) == 8 and result.startswith('2'):

            # check if result is int (does not contain letters)
            try:
                int(result)

            # if result is not number, add to error folder and continue
            except ValueError:
                img.save(ErrorPdfPath + 'noName-' + currentDate + '.pdf')
                os.rename(path + files, ErrorImagePath + files)

                continue

            # if result is valid, save to processed folder and add to xlsx sheet
            img.save(ProcessedPdfPath + result + '.pdf')
            os.rename(path + files, ProcessedImagePath + files)

            # Define cell address
            ws['A' + str(lastRow)] = result
            # increment row number
            lastRow = lastRow + 1

        # If result is Not 8 characters long and/or does not start with int 2
        else:

            img.save(ErrorPdfPath + result + '.pdf')
            os.rename(path + files, ErrorImagePath + files)

    # if result is not detected, add to error folder and continue
    except ValueError:
        img.save(ErrorPdfPath + 'noName-' + currentDate + '.pdf')
        os.rename(path + files, ErrorImagePath + files)
        continue

# Save and close xlsx sheet
wb.save(XslxPath)
wb.close()